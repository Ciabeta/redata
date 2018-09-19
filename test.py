# -*- coding: utf-8 -*-
import time, xlrd
from datetime import timedelta, datetime
from openpyxl import load_workbook
command1 = 'l201600800237\n联系QQ123456'
command2 = 'f201600800237\n联系QQ23456'
command6 = 'l一个棒棒糖\n大概在三食堂四楼，联系qq123456'
command3 = 'lost'
command4 = 'found'
command5 = '19l3s'

log_of_data = '/root/data/find_log/'
#open excel
def open_log_data(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)
        return 0
#有数据返回数据，没有数据返回0
#list: main, sub, row
def get_xlsxdata(file_name):
    data = open_log_data(file_name)
    if data is not None:
        table = data.sheet_by_name('Sheet1')
        #colnames is a matrix of the specific row
        colnames = table.row_values(0)
        list =[]
        #read a table row by row
        for rownum in range(0, table.nrows):
            #get row
            row = table.row_values(rownum)
            
            if row:
                
                app = []
                #read a row col by col
                for i in range(len(colnames)):
                    app.append(row[i])
                app.append(rownum)
                
                list.append(app)
        return list
    else:
        return 0
#log_data: ['main info', 'subinfo','solved'/'unsolved']
def put_xlsx_data(file_name, data_log):
    rexcel = open_log_data(file_name)
    rows = rexcel.sheets()[0].nrows # 用wlrd提供的方法获得现在已有的行数
    ex = load_workbook(file_name)
    rows += 1
    if ex is not None:
        ws = ex['Sheet1']
        ws.cell(row=rows, column=1).value = data_log[0]
        ws.cell(row=rows, column=2).value = data_log[1]
        ws.cell(row=rows, column=3).value = 'unsolved'
        ex.save(file_name)
        return 1
    else:
        #要写的表不存在，创建新表，然后写
        return 0
#if command 25l2s,25f3s
def if_contain_exact(command, exact):
    nPos = command.index(exact)
    if(nPos >=0):
        return 1
    else:
        return 0
def get_which_day(orbefore):
    if(orbefore == 0):
        return time.strftime("%m_%d")
    elif(orbefore == -1):
        yesterday = datetime.today() + timedelta(-1)
        return yesterday.strftime("%m_%d")
    elif(orbefore == -2):
        lastd = datetime.today() + timedelta(-2)
        return lastd.strftime("%m_%d")
    elif(orbefore == 'd'):
        return time.strftime("%d")
    elif(orbefore == 'd-1'):
        yesterday = datetime.today() + timedelta(-1)
        return yesterday.strftime("%d")
#command: ['19', '2', 'f']
#the second param: 2, row start with 0, don't forget +1
def change_status_solved(command):
    if(command[2] == 'f'):
        absolute_addr = log_of_data + 'found_'
    else:
        absolute_addr = log_of_data + 'lost_'
    if(command[0] == get_which_day('d')):
        absolute_addr += time.strftime("%m_%d.xlsx")
    elif(command[0] == get_which_day('d-1')):
        absolute_addr += get_which_day(-1) + '.xlsx'
    else:
        return '修改失败，日期填写错误。'
    ex = load_workbook(absolute_addr)
    if ex is not None:
        ws = ex['Sheet1']
        ws.cell(row=(int(command[1])+1), column=3).value = 'solved'
        ex.save(absolute_addr)
        return '标记为已完成' 
    else:
        return '修改失败，日期填写错误。'

#find info and return data
#if lost or found command
def list_recent_info(command):
    if(command == 'lost'):
        absolute_addr = log_of_data+time.strftime("lost_%m_%d.xlsx")
        lost_list = get_xlsxdata(absolute_addr)
        if(lost_list != 0):
            i = 1
            for oneday in lost_list:
                if(oneday[2] != 'solved'):
                    print "19-%d:%s\n\t%s" % (oneday[3], oneday[0].encode("utf-8"), oneday[1].encode("utf-8"))
    elif(command == 'found'):
        absolute_addr = log_of_data+time.strftime("found_%m_%d.xlsx")
        found_list = get_xlsxdata(absolute_addr)
        if(found_list != 0):
            i = 1
            for oneday in found_list:
                if(oneday[2] != 'solved'):
                    print "19-%d:%s\n\t%s" % (i, oneday[0].encode("utf-8"), oneday[1].encode("utf-8"))
    elif(command.startswith('l')):
         #后续信息插入lost表，如果表不存在，则创建
        command = command.strip("l")
        command = command.split('\n',1)
        absolute_addr = log_of_data+time.strftime("lost_%m_%d.xlsx")
        if(put_xlsx_data(absolute_addr, command)):
            return "写入lost成功,近期请关注群内消息balabala"
    elif(command.startswith('f')):
        #后续信息插入found表，如果表不存在，则创建
        command = command.strip("l")
        command = command.split('\n',1)
        absolute_addr = log_of_data+time.strftime("found_%m_%d.xlsx")
        if(put_xlsx_data(absolute_addr, command)):
            return "写入found成功,近期请关注群内消息balabala"
    elif(if_contain_exact(command, 'l') & if_contain_exact(command, 's')):
        command = command.split('l',1)
        command[1] = command[1].strip('s')
        command.append('l')
        return change_status_solved(command)
    else:
        print command
        return '命令格式不正确，请查阅xxxxxxxx'
        
resu = list_recent_info('lost')
if resu is not None:
    print resu
#在主调用函数一定先看有没#,没有直接跳过不判断，有的话把井号去掉。




#if(isinstance(exec_info, list)):
202                 #    bot.SendTo(contact, '是个数组')
203                 #elif(exec_info != None):
204                 #    bot.SendTo(contact, exec_info)
205                 #else:
206                 #    bot.SendTo(contact,'条件都不成立')
207             #return_stuinfo = post_card_info(content)
208             #if(return_stuinfo != 0):
209             #    bot.SendTo(contact, return_stuinfo)