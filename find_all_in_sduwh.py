# -*- coding: utf-8 -*-
import  xdrlib ,sys
import xlrd
import time, xlrd
from datetime import timedelta, datetime
from openpyxl import load_workbook
#absolute index of files
jd15 = '/root/.qqbot-tmp/plugins/15jdy.xlsx'
jd16 = '/root/.qqbot-tmp/plugins/16jdy.xlsx'
jd17 = '/root/.qqbot-tmp/plugins/17jdy.xlsx'
log_of_data = '/root/data/find_log/'
#open excel
def open_log_data(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)
        return 0
#有数据返回数据，没有数据返回0
#list: main, sub, row
def get_xlsxdata(file_name):
    data = open_log_data(file_name)
    if data is not None:
        table = data.sheet_by_name('Sheet1')
        #colnames is a matrix of the specific row
        colnames = table.row_values(0)
        list =[]
        #read a table row by row
        for rownum in range(0, table.nrows):
            #get row
            row = table.row_values(rownum)
            
            if row:
                
                app = []
                #read a row col by col
                for i in range(len(colnames)):
                    app.append(row[i])
                app.append(rownum)
                
                list.append(app)
        return list
    else:
        return 0
#log_data: ['main info', 'subinfo','solved'/'unsolved']
def put_xlsx_data(file_name, data_log):
    rexcel = open_log_data(file_name)
    rows = rexcel.sheets()[0].nrows # 用wlrd提供的方法获得现在已有的行数
    ex = load_workbook(file_name)
    rows += 1
    if ex is not None:
        ws = ex['Sheet1']
        ws.cell(row=rows, column=1).value = data_log[0]
        ws.cell(row=rows, column=2).value = data_log[1]
        ws.cell(row=rows, column=3).value = 'unsolved'
        ex.save(file_name)
        return 1
    else:
        #要写的表不存在，创建新表，然后写
        return 0
#if command 25l2s,25f3s
def if_contain_exact(command, exact):
    nPos = command.index(exact)
    if(nPos >=0):
        return 1
    else:
        return 0
def get_which_day(orbefore):
    if(orbefore == 0):
        return time.strftime("%m_%d")
    elif(orbefore == -1):
        yesterday = datetime.today() + timedelta(-1)
        return yesterday.strftime("%m_%d")
    elif(orbefore == -2):
        lastd = datetime.today() + timedelta(-2)
        return lastd.strftime("%m_%d")
    elif(orbefore == 'd'):
        return time.strftime("%d")
    elif(orbefore == 'd-1'):
        yesterday = datetime.today() + timedelta(-1)
        return yesterday.strftime("%d")
#command: ['19', '2', 'f']
#the second param: 2, row start with 0, don't forget +1
def change_status_solved(command):
    if(command[2] == 'f'):
        absolute_addr = log_of_data + 'found_'
    else:
        absolute_addr = log_of_data + 'lost_'
    if(command[0] == get_which_day('d')):
        absolute_addr += time.strftime("%m_%d.xlsx")
    elif(command[0] == get_which_day('d-1')):
        absolute_addr += get_which_day(-1) + '.xlsx'
    else:
        return '修改失败，日期填写错误。'
    ex = load_workbook(absolute_addr)
    if ex is not None:
        ws = ex['Sheet1']
        ws.cell(row=(int(command[1])+1), column=3).value = 'solved'
        ex.save(absolute_addr)
        return '标记为已完成' 
    else:
        return '修改失败，日期填写错误。'

#find info and return data
#if lost or found command
def handle_command(command):
    if(command == 'lost'):
        absolute_addr = log_of_data+time.strftime("lost_%m_%d.xlsx")
        lost_list = get_xlsxdata(absolute_addr)
        if(lost_list != 0):
            i = 1
            for oneday in lost_list:
                if(oneday[2] != 'solved'):
                    print "19-%d:%s\n\t%s" % (oneday[3], oneday[0].encode("utf-8"), oneday[1].encode("utf-8"))
    elif(command == 'found'):
        absolute_addr = log_of_data+time.strftime("found_%m_%d.xlsx")
        found_list = get_xlsxdata(absolute_addr)
        if(found_list != 0):
            i = 1
            for oneday in found_list:
                if(oneday[2] != 'solved'):
                    print "19-%d:%s\n\t%s" % (i, oneday[0].encode("utf-8"), oneday[1].encode("utf-8"))
    elif(command.startswith('l')):
         #后续信息插入lost表，如果表不存在，则创建
        command = command.strip("l")
        command = command.split('\n',1)
        absolute_addr = log_of_data+time.strftime("lost_%m_%d.xlsx")
        if(put_xlsx_data(absolute_addr, command)):
            return "写入lost成功,近期请关注群内消息balabala"
    elif(command.startswith('f')):
        #后续信息插入found表，如果表不存在，则创建
        command = command.strip("l")
        command = command.split('\n',1)
        absolute_addr = log_of_data+time.strftime("found_%m_%d.xlsx")
        if(put_xlsx_data(absolute_addr, command)):
            return "写入found成功,近期请关注群内消息balabala"
    elif(if_contain_exact(command, 'l') & if_contain_exact(command, 's')):
        command = command.split('l',1)
        command[1] = command[1].strip('s')
        command.append('l')
        return change_status_solved(command)
    else:
        print command
        return '命令格式不正确，请查阅xxxxxxxx'

#colnameindex：index of file  ，by_name：sheet name
def get_all_studata(file, colnameindex, by_name):
    data = open_log_data(file)
    table = data.sheet_by_name(by_name)
    #colnames is a matrix of the specific row
    colnames = table.row_values(colnameindex)
    list =[]
    #read a table row by row
    for rownum in range(0, table.nrows):
        #get row
        row = table.row_values(rownum)
        if row:
            app = []
            #read a row col by col
            for i in range(len(colnames)):
               app.append(row[i])
            list.append(app)
    return list
def stu_infocheck(card_content):
    id_num = int(card_content)
    if(id_num <= 201500800001):
        return 0
	if(id_num <= 201600800001):
		return jd15
    elif(id_num <= 201700800001):
        return jd16
    elif(id_num <= 201800800001):
        return jd17
    else:
		return 0
def post_card_info(cardcontent):
    if(cardcontent.startswith('#')):
        cardcontent = cardcontent.strip('#')
        file_select = stu_infocheck(cardcontent)
        tables = get_all_studata(file_select, 0, 'Sheet1')
        for row in tables:
	        stuinfo_id = int(cardcontent)
	        if(row[2] == stuinfo_id):
	            return "他是%s的%s噢，请认识的同学通知一下吧！" % (row[0].encode("utf-8"), row[1].encode("utf-8"))
	else:
	    return 0
def onQQMessage(bot, contact, member, content):
    if(contact.ctype == 'group'):
        if(contact.name == '找遍山威Beta'):
            if(contact.startswith('#')):
                contact = contact.strip('#')
                handle_command(contact)
            #return_stuinfo = post_card_info(content)
            #if(return_stuinfo != 0):
            #    bot.SendTo(contact, return_stuinfo)
